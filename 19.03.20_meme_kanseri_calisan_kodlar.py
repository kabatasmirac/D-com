import os
import tkinter as tk
from tkinter import ttk
from PIL import ImageTk
import pydicom
from tkinter import filedialog
import numpy as np
from PIL import *
import PIL
import imageio
from skimage import io
from skimage.filters import threshold_yen
import skimage
from skimage import *
from skimage.morphology import square
from PIL import Image, ImageDraw
import cv2
import PIL.Image
from openpyxl import Workbook
import openpyxl
import pathlib
from PIL import ImageFont, ImageEnhance
from PIL import Image, ImageDraw

class AutoScrollbar(ttk.Scrollbar):
    ''' A scrollbar that hides itself if it's not needed.
        Works only if you use the grid geometry manager '''
    def set(self, lo, hi):
        if float(lo) <= 0.0 and float(hi) >= 1.0:
            self.grid_remove()
        else:
            self.grid()
        ttk.Scrollbar.set(self, lo, hi)
    
        a_degeri,b=ttk.Scrollbar.get(self)

        return a_degeri
        
    def pack(self, **kw):
        raise tk.TclError('Cannot use pack with this widget')

    def place(self, **kw):
        raise tk.TclError('Cannot use place with this widget')


       
class Zoom(ttk.Frame):
    ''' Simple zoom with mouse wheel '''
    

    def __init__(self):
        ''' Initialize the main Frame '''
        self.controls = ttk.Frame(root1,width=400, height = 300)
        ttk.Label(self.controls, text='---Doku Açıklamaları---',font=('arial 18')).grid(row=25,column=0,columnspan=2,sticky="W",padx=5,pady=5)
        scale=1.0
        file = pathlib.Path("Etiketlenen_Lezyon_Goruntuleri") 

        if file.exists()==False:
            os.mkdir("Etiketlenen_Lezyon_Goruntuleri")
        file = pathlib.Path("Dicom_images") 
        if file.exists()==False:
            os.mkdir("Dicom_images")
            
        self.temp_labelID=None
        self.rect=None
        self.image=None
        
        self.master=ttk.Frame(root1,width=350, height =400)
        self.controls.grid(row=0,column=4,sticky="E",padx=10)
        
        self.attr1=ttk.Label(self.controls,text="Tür")
        self.attr2=ttk.Label(self.controls,text="Tür açıklaması")
        self.attr3=ttk.Label(self.controls,text="Kontür")
        self.attr4=ttk.Label(self.controls,text="Cilt Kalınlaşması")
        self.attr5=ttk.Label(self.controls,text="Lenf Nodu")
        self.attr6=ttk.Label(self.controls,text="Eşlik Eden Kalsifikasyon")
        self.attr7=ttk.Label(self.controls,text="attr7")
        self.attr8=ttk.Label(self.controls,text="attr8")
        self.attr9=ttk.Label(self.controls,text="attr9")
        self.attr10=ttk.Label(self.controls,text="attr10")
        self.attr11=ttk.Label(self.controls,text="attr11")
        self.attr12=ttk.Label(self.controls,text="attr12")
        
        self.ent_attr1=tk.StringVar(self.controls)
        optionslist1=["Benign ", "Malignant"]
        option1 = ttk.OptionMenu(self.controls, self.ent_attr1,optionslist1[0] ,*optionslist1)
        option1.grid(row=0,column=1, padx=5, pady=5)
        self.ent_attr2=ttk.Entry(self.controls)
        self.ent_attr3 = tk.IntVar(self.controls)

        R1 = ttk.Radiobutton(self.controls, text = "Silik", 
                         variable = self.ent_attr3, value = 1)
        R2 = ttk.Radiobutton(self.controls, text = "Silik Değil", 
                         variable = self.ent_attr3, value = 0)
        
        
        R1.grid(row = 3, column = 0, padx = 10)
        R2.grid(row = 3, column = 1, padx = 10)
          
        self.ent_attr4 = tk.IntVar(self.controls)

        R3 = ttk.Radiobutton(self.controls, text = "Yok", 
                         variable = self.ent_attr4, value = 1)
        R4 = ttk.Radiobutton(self.controls, text = "Var", 
                         variable = self.ent_attr4, value = 0)
        R3.grid(row = 5, column = 0, padx = 10)
        R4.grid(row = 5, column = 1, padx = 10)
          
        self.ent_attr5 = tk.IntVar(self.controls)

        R5 = ttk.Radiobutton(self.controls, text = "Yok", 
                         variable = self.ent_attr5, value = 1)
        R6 = ttk.Radiobutton(self.controls, text = "Var", 
                         variable = self.ent_attr5, value = 0)
        
        
        R5.grid(row = 7, column = 0, padx = 10)
        R6.grid(row = 7, column = 1, padx = 10)
        self.ent_attr6 = tk.IntVar(self.controls)

        R7 = ttk.Radiobutton(self.controls, text = "Yok", 
                         variable = self.ent_attr6, value = 1)
        R8 = ttk.Radiobutton(self.controls, text = "Var", 
                         variable = self.ent_attr6, value = 0)
        
        
        R7.grid(row = 9, column = 0, padx = 10)
        R8.grid(row = 9, column = 1, padx = 10)
        self.ent_attr7=ttk.Entry(self.controls)
        self.ent_attr8=ttk.Entry(self.controls)
        self.ent_attr9=ttk.Entry(self.controls)
        self.ent_attr10=ttk.Entry(self.controls)
        self.ent_attr11=ttk.Entry(self.controls)
        self.ent_attr12=ttk.Entry(self.controls)
        
        
        self.attr1.grid(row=0,column=0, padx=5, pady=5)
        self.attr2.grid(row=1,column=0, padx=5, pady=5)
        self.ent_attr2.grid(row=1,column=1, padx=5, pady=5)
        
        self.attr3.grid(row=2,column=0,columnspan=3, padx=5, pady=5)
        self.attr4.grid(row=4,column=0,columnspan=3, padx=5, pady=5)
        self.attr5.grid(row=6,column=0,columnspan=3, padx=5, pady=5)
        self.attr6.grid(row=8,column=0,columnspan=3, padx=5, pady=5)
        self.attr7.grid(row=10,column=0, padx=5, pady=5)
        self.ent_attr7.grid(row=10,column=1, padx=5, pady=5)
        
        self.attr8.grid(row=11,column=0, padx=5, pady=5)
        self.ent_attr8.grid(row=11,column=1, padx=5, pady=5)
        
        self.attr9.grid(row=12,column=0, padx=5, pady=5)
        self.ent_attr9.grid(row=12,column=1, padx=5, pady=5)
        
        self.attr10.grid(row=13,column=0, padx=5, pady=5)
        self.ent_attr10.grid(row=13,column=1, padx=5, pady=5)
        
        self.attr11.grid(row=14,column=0, padx=5, pady=5)
        self.ent_attr11.grid(row=14,column=1, padx=5, pady=5)
        
        self.attr12.grid(row=15,column=0, padx=5, pady=5)
        self.ent_attr12.grid(row=15,column=1, padx=5, pady=5)
        
        attrbtn=ttk.Button(self.controls,text="Save!",command=self.addsheet)
        attrbtn.grid(row=18,column=1,columnspan=2,padx=5,pady=5)
        
        
        self.master.grid(row=0,column=0,sticky="WESN")
        ttk.Frame.__init__(self,)
        vbar = AutoScrollbar(self.master, orient='vertical')
        hbar = AutoScrollbar(self.master, orient='horizontal')
        vbar.grid(row=0, column=1, sticky='ns')
        hbar.grid(row=1, column=0, sticky='we')
        self.open_image()
        
        if self.image:
            self.canvas = tk.Canvas(self.master,width=scale*self.width,height=scale*self.height,
                                    xscrollcommand=hbar.set, yscrollcommand=vbar.set)
            
            self.canvas.grid(row=0, column=0, sticky='we')
            vbar.configure(command=self.canvas.yview)  # bind scrollbars to the canvas
            hbar.configure(command=self.canvas.xview)
            # Make the canvas expandable
            self.master.rowconfigure(0, weight=1)
            self.master.columnconfigure(0, weight=1)
            # Bind events to the Canvas
            self.canvas.bind('<Button-3>', self.move_from)
            self.canvas.bind('<B3-Motion>', self.move_to)
            self.canvas.bind('<Control-MouseWheel>', self.wheel)  # with Windows and MacOS, but not Linux
            self.canvas.bind("<ButtonPress-1>", self.on_button_press)
            self.canvas.bind("<B1-Motion>", self.on_move_press)
#            self.canvas.bind("<ButtonRelease-1>", self.on_button_release)
            self.canvas.bind("<ButtonRelease-1>", self.crop)
            self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
            self.canvas.bind_all("<Control-z>",self.geri1)
            # Show image and plot some random test rectangles on the canvas
            self.imscale = 1.0
            self.imageid = None
            self.imscale1 = 1.0
            self.imageid1 = None
            self.delta = 1.25
            width, height = self.image.size
            self.show_image()
            self.canvas.configure(scrollregion=self.canvas.bbox('all'))
            scale=1.0

            self.rect = None
            self.start_x = None
            self.start_y = None
            self.curX=None
            self.curY=None
            self.scale=None
            self.scale1=None
            self.fit=None
            self.imageid1=None
            self.canvas1=None
            self.dolu=None
            self.save_path=None
            self.positions = []
        else:
            print("Dosya bulunamadı..\nNo file..")

    def addsheet(self):
        self.sheetname="MLPCAREML.xlsx"
        self.koordinatlar="LEZYONCOORDINATE.xlsx"
   
        self.mouseclick=self.mouseclick+1
        if os.path.exists(self.sheetname) == False:
            book = Workbook()
            sheet = book.active
            sheet["A1"]='DICOM_ID'
            sheet["B1"]='LABEL_ID'
            sheet["C1"]='TUMOR_TYPE'
            sheet["D1"]='SHAPE'
            sheet["E1"]='DENSITY'
            sheet["F1"]='KALINLIK'
            sheet["G1"]='ATTR5'
            sheet["H1"]='ATTR6'
            sheet["I1"]='ATTR7'
            sheet["J1"]='ATTR8'
            sheet["K1"]='ATTR9'
            sheet["L1"]='ATTR10'
            sheet["M1"]='ATTR11'
            sheet["N1"]='ATTR12'
            book.save(self.sheetname)
            print("Dosya oluşturuldu..\nFile created..")
            
        
        if self.mouseclick != 0:
            
            sheet1 = openpyxl.load_workbook(filename = self.sheetname)
            sheet=sheet1.active
            name1=self.ent_attr1.get()
            name2=self.ent_attr2.get()
            name3=self.ent_attr3.get()
            name4=self.ent_attr4.get()
            name5=self.ent_attr5.get()
            name6=self.ent_attr6.get()
            name7=self.ent_attr7.get()
            name8=self.ent_attr8.get()
            name9=self.ent_attr9.get()
            name10=self.ent_attr10.get()
            name11=self.ent_attr11.get()
            name12=self.ent_attr12.get()
            
            max_rowi=sheet.max_row
            sheet["A"+str(max_rowi+1)]=self.image_name
            sheet["B"+str(max_rowi+1)]=self.tumor
            sheet["C"+str(max_rowi+1)]=name1
            sheet["D"+str(max_rowi+1)]=name2
            sheet["E"+str(max_rowi+1)]=name3
            sheet["F"+str(max_rowi+1)]=name4
            sheet["G"+str(max_rowi+1)]=name5
            sheet["H"+str(max_rowi+1)]=name6
            sheet["I"+str(max_rowi+1)]=name7
            
            sheet["J"+str(max_rowi+1)]=name8
            sheet["K"+str(max_rowi+1)]=name9
            sheet["L"+str(max_rowi+1)]=name10
            sheet["M"+str(max_rowi+1)]=name11
            sheet["N"+str(max_rowi+1)]=name12
            if self.temp_labelID!=self.tumor:
                sheet1.save(self.sheetname)
                print("Etiketlendi..\nTagged..")
            else:
                print("Aynı etiketi birden fazla kullanamazsınız..\nYou cannot use the same tag more than once..")
            
            self.temp_labelID=self.tumor
            
        else:
            print("Etiketleme yapılamadı..\nTagging failed..")
            
        self.koordinatlar="LEZYONCOORDINATE.xlsx"
        if os.path.exists(self.koordinatlar) == False:
            book = Workbook()
            sheet = book.active
            sheet["A1"]='DICOM_ID'
            sheet["B1"]='LABEL_ID'
            sheet["C1"]='START_X_COORDINATE'
            sheet["D1"]='START_Y_COORDINATE'
            sheet["E1"]='END_X_COORDINATE'
            sheet["F1"]='END_Y_COORDINATE'
            
            book.save(self.koordinatlar)
            print("Dosya oluşturuldu..\nFile created..")
            
        
        if self.mouseclick != 0:
            print(self.mouseclick)
            sheet1 = openpyxl.load_workbook(filename = self.koordinatlar)
            sheet=sheet1.active
            max_rowi=sheet.max_row
            sheet["A"+str(max_rowi+1)]=self.image_name
            sheet["B"+str(max_rowi+1)]=self.tumor
            sheet["C"+str(max_rowi+1)]=self.start_x
            sheet["D"+str(max_rowi+1)]=self.start_y
            sheet["E"+str(max_rowi+1)]=self.curX
            sheet["F"+str(max_rowi+1)]=self.curY
            sheet1.save(self.koordinatlar)
            print("Kaydedildi..\nRecorded..")
    global cord_list
    cord_list=list()
    global del_list
    del_list=list()
    def poly(self):
        self.canvas.bind("<Button-1>", self.draw_line)
        self.canvas.bind("<Button-3>", self.convert_poligon)        
        self.canvas.bind_all("<Control-z>",self.line_geri_al)
       
        self.canvas.bind("<ButtonRelease-1>", self.paspaspas)
    def recta(self):
        self.canvas.bind('<Button-3>', self.move_from)
        self.canvas.bind('<B3-Motion>', self.move_to)
        self.canvas.bind('<Control-MouseWheel>', self.wheel)  # with Windows and MacOS, but not Linux
        self.canvas.bind("<ButtonPress-1>", self.on_button_press)
        self.canvas.bind("<B1-Motion>", self.on_move_press)
        self.canvas.bind("<ButtonRelease-1>", self.crop)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
    global geri_alinan
    geri_alinan=list()
    global ctrlz
    ctrlz=list()

    def line_geri_al(self,event):
        if self.controla == 0:
            a=ctrlz.pop()
            self.canvas.delete(a)
            geri_alinan.append(a)
            a=ctrlz.pop()
            self.canvas.delete(a)
            geri_alinan.append(a)
            cord_list.pop()
            ass = self.canvas.create_line(cord_list[len(cord_list)-1][0],cord_list[len(cord_list)-1][1],cord_list[0][0],cord_list[0][1],fill='white',width=5)
            ctrlz.append(ass)
            del_list[0]=ass
            self.controla = 1
        else:
            pass
    def paspaspas(self, event):
        pass
    global silhepsini
    silhepsini = list()
    def draw_line(self, event):
        self.controla=0 
        self.click_number=0
        self.start_x,self.start_y
        self.click_number=len(cord_list)
        if self.click_number==0:
            self.start_x=self.canvas.canvasx(event.x)
            self.start_y=self.canvas.canvasy(event.y)
            self.click_number=1
            cord_list.append([self.start_x,self.start_y])
            
        else:
            x2=self.canvas.canvasx(event.x)
            y2=self.canvas.canvasy(event.y)
            ass=self.canvas.create_line(cord_list[len(cord_list)-1][0],cord_list[len(cord_list)-1][1],x2,y2,fill='red',width=3)
            ctrlz.append(ass)
            silhepsini.append(ass)
            self.click_number=len(cord_list)
            cord_list.append([x2,y2])
            self.controla = 0
        if (self.click_number>2 ):
            
            a = self.canvas.create_line(cord_list[len(cord_list)-1][0],cord_list[len(cord_list)-1][1],cord_list[0][0],cord_list[0][1],fill='red',width=3)
            ctrlz.append(a)
            silhepsini.append(a)
            del_list.append(a)
            self.controla = 0
        if(len(del_list)==2):
            self.canvas.delete(del_list[0])
            del_list.pop(0)
        
    global points
    points=list()
    global polygon
    polygon = list()
    def hepsinisil(self,i):
        self.canvas.delete(i)
    def convert_poligon(self,event):   
        self.mask=None
        for i in silhepsini:
            self.hepsinisil(i)
        points=np.array(cord_list,dtype=np.int32)
        im=PIL.Image.open(os.getcwd()+"\\Dicom_images\\"+self.image_name+'test.png').convert("RGB")
        ###^!!!!! RGB iken çalışıyor asla değiştirme haberin olsun
        img=np.asarray(im)
        pts = points
        ## (1) Crop the bounding rect
        rect = cv2.boundingRect(pts)
        x,y,w,h = rect
        self.start_x=x
        self.start_y=y
        self.curX=x+w
        self.curY=y+h
        self.cropped_img = img[y:y+h, x:x+w].copy()
        ## (2) make mask
        pts = pts - pts.min(axis=0)
        self.mask = np.zeros(self.cropped_img.shape[:2], np.uint8)
        cv2.drawContours(self.mask, [pts], -1, (255, 255, 255), -1, cv2.LINE_AA)
        cv2.imwrite("croped.png", self.cropped_img )
        cv2.imwrite("mask.png", self.mask)
        ## (3) do bit-op
        area = x,y,x+w,y+h
        self.cropped_img1=self.photo.crop(area)
        self.cropped_img1.save("cropp.png","PNG")
        self.cropped_img1=cv2.imread("cropp.png")
        self.dst = cv2.bitwise_and(self.cropped_img1, self.cropped_img1, mask=self.mask) # gerçekten poligon alanın treshol yapmak istiyorsan kullan
        #self.dst=cv2.imread("mask.png") # poligonun tresholda ihtiyacı yok diyorsan kullan
        cv2.imwrite("dst.png", self.dst)##bu kısımla ilgileneceğiz
        ## (4) add the white background
        bg = np.ones_like(self.cropped_img, np.uint8)*255
        self.pathi=os.getcwd()+"\\Etiketlenen_Lezyon_Goruntuleri\\"
        self.tumor = self.tumor+1
        print(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"))
        self.bisilerbisiler1()
        self.bolukporcuk1()
        self.pixelesikleme()
        self.save_image()
        self.create()
        
        del cord_list[:]
        
        del del_list[:]
    def bisilerbisiler1(self):
        
        h,w=self.cropped_img1.shape[:2]        
        self.dst=PIL.Image.open("dst.png")
        if self.scale:
            self.new_size=int(w/self.imscale),int(h/self.imscale)
            self.dst=self.dst.resize(self.new_size) 
            self.dst.save(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"), "PNG")
        else:
            self.dst.save(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"), "PNG")
#
##        HİSTOGRAM EŞİKLEME 
#        img5 = cv2.imread(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"),0)
#        img5 = cv2.equalizeHist(img5)
#        img5 = cv2.equalizeHist(img5)
#        cv2.imwrite(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"),img5)
        
        ##Gaussian filtre
##        
#        img = cv2.imread(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"),0)
#        blur = cv2.GaussianBlur(img,(125,125),0)
#        cv2.imwrite(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"),blur)
    def bolukporcuk1(self):  
        self.img=skimage.io.imread(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"),as_gray=True)
#        self.img=cv2.imread(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"))
#        self.gray = cv2.cvtColor(self.img,cv2.COLOR_BGR2GRAY)
        thr = threshold_yen(self.img)
        thr = thr - np.float64(0.14)
        binary = self.img <= thr
        bins = np.zeros([self.img.shape[0],self.img.shape[1]], dtype=np.uint8)   
        self.threshold = np.zeros([self.img.shape[0],self.img.shape[1]], dtype=np.uint8)

        for i in range (self.img.shape[1]):
            for j in range (self.img.shape[0]):
                if binary[j][i] == True:
                    self.threshold[j][i] = 0
                else:
                    self.threshold[j][i] = 255
                    
#        kernel = np.ones((3,3), np.uint8) 
#        img_dilation = cv2.dilate(self.threshold, kernel, iterations=1) 
#        cv2.imwrite('Dilation.png', img_dilation) 
#        retval, self.threshold = cv2.threshold(self.gray, 120, 255, cv2.THRESH_BINARY)
        cv2.imwrite("threshhold.png",self.threshold)
        
        self.dilation_img=skimage.morphology.erosion(self.threshold)         
        imageio.imwrite('image_name.png',self.dilation_img)
        self.threshold_img=PIL.Image.open("image_name.png")
        
        
    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/120)),"units")
        
    def open_image(self):
        
        self.mouseclick=0
        self.tumor=0
        self.labelnum=list()
        global image_path
        image_path = filedialog.askopenfilename()
        
        if len(image_path) > 0:
            ds = pydicom.dcmread(image_path)
            self.shape = ds.pixel_array.shape
            image_2d = ds.pixel_array
            image_2d_scaled = (np.maximum(image_2d,0) / image_2d.max()) * 255.0
            self.image_name=image_path.split('/')[len(image_path.split('/'))-1].replace(".dcm","")
            os.mkdir(os.getcwd()+"\\Etiketlenen_Lezyon_Goruntuleri\\"+self.image_name)
            image_2d_scaled = np.uint8(image_2d_scaled)
            im=PIL.Image.fromarray(image_2d_scaled)
            im.save(os.getcwd()+"\\Dicom_images\\"+self.image_name+'test.png')
            self.image = PIL.Image.open(os.getcwd()+"\\Dicom_images\\"+self.image_name+'test.png')
            self.width, self.height = self.image.size
            self.blank_img = PIL.Image.new('RGB', (self.width,self.height),"black")
            self.blank_img.save("blank_img.tiff", "TIFF")
        else:
            print("Görüntü bulunamadı veya Dicom formatında değil!!!\nImage not found or not in Dicom format!!!")

    def on_button_press(self, event):
        
        self.tumor=self.tumor+1
        self.canvas=event.widget
        # save mouse drag start position
        self.start_x = self.canvas.canvasx(event.x)
        self.start_y = self.canvas.canvasy(event.y)

        if not self.rect:
            self.rect = self.canvas.create_rectangle(self.start_x, self.start_y, 1, 1, outline="red")
            
    def on_move_press(self, event):
        self.curX = self.canvas.canvasx(event.x)
        self.curY = self.canvas.canvasy(event.y)
        
        self.canvas.coords(self.rect, self.start_x, self.start_y, self.curX, self.curY)
        
    def on_button_release(self, event):
        pass
    
    def move_from(self, event):
        ''' Remember previous coordinates for scrolling with the mouse '''
        self.canvas.scan_mark(event.x, event.y)

    def move_to(self, event):
        ''' Drag (move) canvas to the new position '''
        self.canvas.scan_dragto(event.x, event.y, gain=1)
    
    def move_from1(self, event):
        ''' Remember previous coordinates for scrolling with the mouse '''
        self.canvas1.scan_mark(event.x, event.y)

    def move_to1(self, event):
        ''' Drag (move) canvas to the new position '''
        self.canvas1.scan_dragto(event.x, event.y, gain=1)   

    def wheel(self,event):
        
        ''' Zoom with mouse wheel '''
        global scale,imscale,width,height
        scale=1.0
        # Respond to Windows (event.delta) wheel event
        if event.delta == 120:
            self.imscale *= self.delta
            scale *= self.delta
            print("self.imscale",self.imscale)
            self.scale=scale
        if event.delta == -120:
            self.imscale /= self.delta
            scale/=self.delta
            self.scale=scale
            print("self.imscale",self.imscale)
        # Rescale all canvas objects
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        self.scalemmm=round(self.imscale)
        self.canvas.scale('all', x, y,scale,scale)
        self.show_image()
        self.canvas.configure(scrollregion=self.canvas.bbox('all'))   
            
    
    def wheel1(self,event):
        ''' Zoom with mouse wheel '''
        global scale1,imscale1,width1,height1
        scale1=1.0
        # Respond to Windows (event.delta) wheel event
        if event.delta == 120:
            self.imscale1 *= self.delta
            scale1 *= self.delta
            print("self.imscale1",self.imscale1)
            self.scale1=scale1
        if event.delta == -120:
            self.imscale1 /= self.delta
            scale1/=self.delta
            self.scale1=scale1
            print("self.imscale1",self.imscale1)
        # Rescale all canvas objects
        x = self.canvas1.canvasx(event.x)
        y = self.canvas1.canvasy(event.y)
        
        z=ttk.Scrollbar.get()
        self.canvas1.scale('all', x, y,scale1,scale1)
        self.show_image1()
        self.canvas1.configure(scrollregion=self.canvas1.bbox('all'))
        
    def crop(self,event):
        global area,position
        
        if self.start_x>self.curX and self.start_y<self.curY:
            self.area=(self.curX,self.start_y,self.start_x,self.curY)
        elif self.start_x<self.curX and self.start_y>self.curY:
            self.area=(self.start_x,self.curY,self.curX,self.start_y)
        elif self.start_x>self.curX and self.start_y>self.curY:
            self.area=(self.curX,self.curY,self.start_x,self.start_y)
        else:
            self.area = (self.start_x,self.start_y,self.curX,self.curY)
        self.cropped_img = self.photo.crop(self.area)
        self.pathi=os.getcwd()+"\\Etiketlenen_Lezyon_Goruntuleri\\"
        print(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"))

        self.bisilerbisiler()
        self.bolukporcuk()
        self.pixelesikleme()
        self.save_image()
        self.create()
        
    def bisilerbisiler(self):
        w,h=self.cropped_img.size
        if self.scale:
            self.new_size=int(w/self.imscale),int(h/self.imscale)
            self.cropped_img1=self.cropped_img.resize(self.new_size) 
            self.cropped_img1.save(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"), "PNG")
        else:
            self.cropped_img.save(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"), "PNG")
    
    def bolukporcuk(self):   
        self.img=skimage.io.imread(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"),as_gray=True)
        thr = threshold_yen(self.img)
        binary = self.img <= thr
        bins = np.zeros([self.img.shape[0],self.img.shape[1]], dtype=np.uint8)   
        self.threshold = np.zeros([self.img.shape[0],self.img.shape[1]], dtype=np.uint8)

        for i in range (self.img.shape[1]):
            for j in range (self.img.shape[0]):
                if binary[j][i] == True:
                    self.threshold[j][i] = 0
                else:
                    self.threshold[j][i] = 255
                    
        cv2.imwrite("threshhold.png",self.threshold)
        self.dilation_img=skimage.morphology.erosion(self.threshold)         
        imageio.imwrite('image_name.png',self.dilation_img)
        self.threshold_img=PIL.Image.open("image_name.png")
#########PİXEL EŞİTLEMEYİ BURDA YAPIYOR############      
    def pixelesikleme(self):
        if self.scale:
            if self.start_x>self.curX and self.start_y<self.curY:
                self.position=(int(self.curX/self.imscale),int(self.start_y/scale),
                               int(self.curX/self.imscale+self.new_size[0]),int(self.start_y/self.imscale+self.new_size[1]))
            elif self.start_x<self.curX and self.start_y>self.curY:
                self.position=(int(self.start_x/self.imscale),int(self.curY/self.imscale),
                               int(self.start_x/self.imscale+self.new_size[0]),int(self.curY/self.imscale+self.new_size[1]))
            elif self.start_x>self.curX and self.start_y>self.curY:
                self.position=(int(self.curX/self.imscale),int(self.curY/self.imscale),
                               int(self.curX/self.imscale+self.new_size[0]),int(self.curY/self.imscale+self.new_size[1]))
            else:
                self.position=(int(self.start_x/self.imscale),int(self.start_y/self.imscale),
                           int(self.start_x/self.imscale+self.new_size[0]),int(self.start_y/self.imscale+self.new_size[1]))
        else:
            if self.start_x>self.curX and self.start_y<self.curY:
                self.position=(int(self.curX),int(self.start_y),
                               int(self.start_x),int(self.curY))
            elif self.start_x<self.curX and self.start_y>self.curY:
                self.position=(int(self.start_x),int(self.curY),
                               int(self.curX),int(self.start_y))
            elif self.start_x>self.curX and self.start_y>self.curY:
                self.position=(int(self.curX),int(self.curY),
                               int(self.start_x),int(self.start_y))
            else:
                self.position=(int(self.start_x),int(self.start_y),int(self.curX),int(self.curY))
                
        self.positions.append(self.position)
        print("position: ",self.position)
        self.blank=PIL.Image.open('blank_img.tiff')
        self.blank.paste(self.threshold_img,self.position)
        #labelları burada çiziyor
        source_img = PIL.Image.open(os.getcwd()+"\\Dicom_images\\"+self.image_name+'test.png').convert("RGBA")
        shape=[(int(self.start_x),int(self.start_y)),(int(self.curX),int(self.curY))]
        draw = ImageDraw.Draw(source_img)
        draw.rectangle(shape, outline="red")
        draw.text((int(self.start_x),int(self.start_y)), str(self.tumor))
        source_img.save(os.getcwd()+"\\Dicom_images\\"+self.image_name+'test.png', "PNG")

        
        
    def show_image(self):
        
        if self.imageid:
            self.canvas.delete(self.imageid)
            self.imageid = None
            self.canvas.imagetk = None  # delete previous image from the canvas
        width, height = self.image.size
        new_size = int(self.imscale * width), int(self.imscale * height)
        self.photo=self.image.resize((new_size))
        ert,fgh=self.photo.size
      
        self.imagetk = ImageTk.PhotoImage(self.photo,master=self.canvas)
        # Use self.text object to set proper coordinates
        self.imageid = self.canvas.create_image(0,0,anchor='nw', image=self.imagetk)
        self.canvas.lower(self.imageid)
        self.canvas.imagetk = self.imagetk  # keep an extra reference to prevent garbage-collection
        
    def save_image(self):
        if self.dolu:
            self.dolu.save('blank_img.tiff',"TIFF")
            self.dolu=None
        else:
            self.blank.save('blank_img.tiff',"TIFF")
            
    def _on_mousewheel1(self, event):
        self.canvas1.yview_scroll(int(-1*(event.delta/120)),"units")        
            
    def create(self):
        
        if self.canvas1:
            self.image1 = PIL.Image.open('blank_img.tiff')
            self.show_image1()
        else:
            self.master1=tk.Tk()
            self.master1.title("Image's Label")
            menu = tk.Menu(self.master1)
            self.master1.config(menu=menu)
            self.fileMenu = tk.Menu(menu)
            self.fileMenu.add_command(label="Save Labeled Image",command=self.save_label)
            self.fileMenu.add_command(label="Geri Al(ctrl+z)",command=self.geri)
            menu.add_cascade(label="File",menu=self.fileMenu)
            
            vbar1 = AutoScrollbar(self.master1, orient='vertical')
            hbar1 = AutoScrollbar(self.master1, orient='horizontal')
            vbar1.grid(row=0, column=1, sticky='ns')
            hbar1.grid(row=1, column=0, sticky='we')
            
            self.image1 = PIL.Image.open('blank_img.tiff')
            self.width, self.height = self.image1.size
            # Create canvas and put image on it
            self.canvas1 = tk.Canvas(self.master1,width=self.width,height=self.height,
                                    xscrollcommand=hbar1.set, yscrollcommand=vbar1.set)
            self.canvas1.grid(row=0, column=0, sticky='nswe')
            vbar1.configure(command=self.canvas1.yview)  # bind scrollbars to the canvas
            hbar1.configure(command=self.canvas1.xview)
            
            self.canvas1.bind('<Control-MouseWheel>', self.wheel1)
            self.canvas1.bind_all("<MouseWheel>", self._on_mousewheel1)
            self.canvas1.bind_all("<Control-z>",self.geri1)
            self.canvas1.bind('<Button-3>', self.move_from1)
            self.canvas1.bind('<B3-Motion>', self.move_to1)
            self.canvas1.bind_all("<Control-s>",self.save_label1)
            # Make the canvas expandable
            self.master1.rowconfigure(0, weight=1)
            self.master1.columnconfigure(0, weight=1)
            self.canvas1.configure(scrollregion=self.canvas1.bbox('all'))
            self.show_image1()
            
    def save_label1(self,event):
      
        index=image_path.rfind("/")
        self.save_path=image_path+ " labels"
        if os.path.exists(self.save_path):
            save_name=image_path[index+1:len(image_path)]+ " label"
            self.image1.save(os.path.join(self.save_path,save_name+".png"))
            print("Etiket Görüntüsü Kaydedildi..\nThe label image has been saved..")
            
        else:
            os.mkdir(self.save_path)
            save_name=image_path[index+1:len(image_path)]+ " label"
            self.image1.save(os.path.join(self.save_path,save_name+".png"))
            
    def save_label(self):
        """from PIL import Image, ImageGrab """
        index=image_path.rfind("/")
        self.save_path=image_path + " labels"

        if os.path.exists(self.save_path) == False:
            os.mkdir(self.save_path)
        save_name=image_path[index+1:len(image_path)]+ " label"
        self.image1.save(os.path.join(self.save_path,save_name+".png"))
        image1 = cv2.imread('blank_img.tiff',cv2.IMREAD_GRAYSCALE) 
        image2 = cv2.imread(self.save_path+'\\'+save_name+".png",cv2.IMREAD_GRAYSCALE) 
        ret,thresh1 = cv2.threshold(image1,5,255,cv2.THRESH_BINARY)
        ret,thresh2 = cv2.threshold(image2,5,255,cv2.THRESH_BINARY)
        dest_or = cv2.bitwise_or(thresh1, thresh2) 
        cv2.imwrite("mrc.png",thresh1)
        cv2.imwrite("msg.png",thresh2)
        cv2.imwrite(self.save_path+'\\'+save_name+".png",dest_or)
        self.img1= PIL.Image.open(self.save_path+'\\'+save_name+".png")
        self.blank.paste(self.img1, (0,0))
        self.image1.save(os.path.join(self.save_path,save_name+".png"))
        print("Etiket Görüntüsü Kaydedildi..\nThe label image has been saved..")
        
    def geri1(self,event):
        global dolu
        a=len(self.positions)-1
        while a>=0:
            self.dolu=PIL.Image.open('blank_img.tiff')
            draw = ImageDraw.Draw(self.dolu)
            draw.rectangle([int(self.positions[a][0]),int(self.positions[a][1]),int(self.positions[a][2]),int(self.positions[a][3])],fill = (0,0,0))
            self.save_image()
            self.create()
            del self.positions[len(self.positions)-1]
            break
        
    def geri(self):
        global dolu
        a=len(self.positions)-1
        while a>=0:
            self.dolu=PIL.Image.open('blank_img.tiff')
            draw = ImageDraw.Draw(self.dolu)
            draw.rectangle([int(self.positions[a][0]),int(self.positions[a][1]),int(self.positions[a][2]),int(self.positions[a][3])],fill = (0,0,0))
            self.save_image()
            self.create()
            del self.positions[len(self.positions)-1]
            break
    
    def show_image1(self):
        if self.imageid1:
         
            self.canvas1.delete(self.imageid1)
            self.imageid1 = None
            self.canvas1.imagetk1 = None # delete previous image from the canvas
            
        width1, height1 = self.image1.size
        new_size1 = int(self.imscale1 * width1), int(self.imscale1 * height1)
        self.photo1=self.image1.resize((new_size1))
        ert1,fgh1=self.photo1.size
        
        self.imagetk1 = ImageTk.PhotoImage(self.photo1,master=self.canvas1)
        self.imageid1 = self.canvas1.create_image(0,0,anchor='nw', image=self.imagetk1)
        self.canvas1.lower(self.imageid1)
        self.canvas1.imagetk = self.imagetk1  # keep an extra reference to prevent garbage-collection
        
    def delete_image(self):
        if self.imageid:
            self.canvas.delete(self.imageid)
            self.imageid = None
            self.canvas.imagetk = None  # delete previous image from the canvas
        if os.path.exists(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png")):
            os.remove(os.path.join(self.pathi,self.image_name,self.image_name+"_LabelID_"+str(self.tumor)+".png"))
        if os.path.exists("image_name.png"):
            os.remove("image_name.png")
        if os.path.exists("blank_img.tiff"):
            os.remove("blank_img.tiff")
        if os.path.exists("test.tiff"):
            self.image.close()
            os.remove("test.tiff")
       
        self.canvas.delete("all")

from tkinter import *   
    
class Window(Frame):
    def __init__(self, master1=None):
        Frame.__init__(self, master1)
        self.master1 = master1
        menu = Menu(self.master1)
        self.master1.config(menu=menu)
        fileMenu = Menu(menu)
        fileMenu.add_command(label="Open Dicom Image",command=self.open_image)
        fileMenu.add_command(label="Delete Current Image",command=self.delete_image)
        fileMenu.add_command(label="Open New Image",command=self.new_image)
        fileMenu.add_command(label="Exit", command=self.master.destroy)
        menu.add_cascade(label="File",menu=fileMenu)

        selectMenu = Menu(menu)
        selectMenu.add_command(label="Poligony",command=self.poly)
        selectMenu.add_command(label="Rectangle",command=self.recta)
        menu.add_cascade(label="Tool", menu=selectMenu)
 
        helpMenu = Menu(menu)
        helpMenu.add_command(label="Use",command=self.yardim)
        helpMenu.add_command(label="About",command=self.bilgi)
        menu.add_cascade(label="Help", menu=helpMenu)
                
    def open_image(self):
        global app
        app=Zoom()
    
    def delete_image(self):
        print("Siliniyor..\nDeleting..")
        app.delete_image()
        print("Silindi..\nDeleted")
        
    def new_image(self):
        app=Zoom()
        
    def poly(self):
        print('Poligon aracı seçildi..\nPolygon tool selected..')
        app.poly()

    def recta(self):
        print('Dörtgen aracı seçildi..\nRectangle tool selected..')
        app.recta()

    def yardim(self):
        pencere = Tk()
        pencere.title("Usage of Program")
        pencere.geometry("200x400")
        uygulama = Frame(pencere)
        uygulama.grid()
        etiket = Label(uygulama,text="File menüsü altındaki\n'Open Dicom Image'\nkomutu ile dicom\nformatında görüntü açılır.\nAna pencerede mouse\nsağ tuşu ile etiketlenecek\nbölge dikdörtgen\niçerisine alınır.\nProgram etiketlemeyi\notomatik olarak \nyapmaktadır.Etiket \noluşturma aşamasında \nYen sınıflandırma yöntemi\n kullanılmaktadır.  \nEtiket sonuçlarının \n kötü çıkmaması için \n dikdörtgeni mümkün \n olduğunca küçük seçiniz.\nEtiket görüntüsünü\n kaydetmek için yeni\n açılan yardımcı \npencerede File menüsünün\naltındaki Save Labeled Image\nkomutuna tıklayınız.\nTool sekmesi altında\ngörünen Poligon\nseçeneğiyle\ndüzgün olmayan ve\ndamar bölgesiyle karışmış\nolan tümörleri tümor\nolmayan bölgelerden\nayırmak için\nseçebilirsiniz.\nPoligon seçimle çalışırken\nmauseunuzun sol tuşuyla\nnoktalar bırakarak seçim\nseçim yapabilirsiniz.\nHer poligon bölgeden\nsonra etiketleme işleminin\ngerçekleştirebilmek için\nyine mauseunuzun sağ\ntuşuna bir kere\ntıklamanız yeterli olacaktır. \nTekrar\nrectangle seçime\ngeçebilmek için aynı\nsekme altındaki\nrectangle seçeneğini\nseçmeniz yeterli\nolacaktır.\n\nPoligon seçim sırasında\nseçimlerin etiketlenebilmesi\niçin son seçimden sonra\nmauseun sol tuşu\nve sağ tuşuna ardışık\nolarak tıklamanız \nyeterli olacaktır.\nPoligon seçim\nsırasında bir kerelik\ngeri alma hakkınız \nolacaktır.\nYanlış seçtiğinizi\ndüşündüğünüz noktayı\nCTRL-Z kombinasyonu\nile her seçim sonrası\nbir kerelik geri alma\nişlemi gerçekleştirebilirsiniz.\n",font="Times 10",height=25,anchor=N,justify='left')
        etiket.grid(padx=10, pady=10)
        pencere.mainloop()
        
    def bilgi(self):
        pencere1 = Tk()
        pencere1.title("About")
        pencere1.geometry("200x300+100+200")
        uygulama1 = Frame(pencere1)
        uygulama1.grid()
        etiket1 = Label(uygulama1,text="MS hastalığı ve benzeri\nbirçok hastalığın derin\nöğrenme yöntemleri ile\nteşhisinde kullanılacak\nverilerin etiketleme\n(veri hazırlama) işlemini\nkolaylaştırmak ve\nyarı otomatik hale\ngetirmek üzere tasarlanmıştır.\nKodlar Python 3.7 versiyonunda\nyazılmıştır.Kullanılan temel\npaketler:Tkinter,\nPILLOW,pydicom paketleridir. \n(Bu yazılım 11.12.2019 tarihinde \nmemekanseri etiketleme programı adıyla\n MLPCARE ARGE HİZMETLERİ bünyesinde \nMirac Kabataş tarafından tamamlanmıştır.\n Emeği geçen herkese \nteşekkürlerimizi sunarız. \nMiraç Kabataş \n- Temel Sağlık Süreçleri Uygulama Geliştirme Uzman Yardımcısı)",height=30,font="Times 10",anchor=N,justify='left')
        etiket1.grid(padx=10, pady=10)
        pencere1.mainloop()

if __name__ == '__main__':
    
    root1 = Tk()
    app1 = Window(root1)
    root1.geometry("700x700+300+10")
    root1.wm_title("Labeling Tool")
    root1.mainloop()